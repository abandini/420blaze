#!/usr/bin/env python3
"""build-edible-data.py — merge the dispensary EDIBLE spreadsheets into the JSON
the /edibles "Edible Decoder" page reads.

Data flow:  cowork keeps /terrasana/*_edibles.xlsx refreshed -> THIS SCRIPT -> data/edible-products.json

The magic step: an edible that names a real cultivar (in `Strain` or `Source Cultivar`)
INHERITS that cultivar's terpene profile from data/strain-terpenes.json — so a flavor-named
live-resin gummy made from Blue Dream gets Blue Dream's terpenes. Edibles that publish their
own terpene COA use those numbers directly. Everything else is profile:"none" (distillate wall).

Usage:  python3 scripts/build-edible-data.py
"""
import json, re, sys
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
SRC = REPO / "terrasana"
FLOWER = REPO / "data" / "strain-terpenes.json"
OUT = REPO / "data" / "edible-products.json"

TERP_KEYS = ["myrcene", "limonene", "caryophyllene", "linalool", "humulene", "apinene", "bpinene", "bisabolol"]

# edible filename stem (minus _edibles) -> flower dispensary key (for shared store metadata)
STORE_KEY = {
    "terrasana_cleveland": "terrasana", "dacut_monroe": "dacut", "rise_cleveland_pearl": "rise",
    "story_cleveland": "story", "urb_cannabis_monroe": "urb", "joyology_monroe": "joyology",
    "puff_monroe": "puff", "pure_monroe": "pure", "klutch_cleveland": "klutch",
    "botanist_solon": "botanist_solon", "landing_cleveland": "landing",
    "amplify_cleveland_hts": "amplify", "forest_lakewood": "forest_lakewood",
    "roam_seven_hills": "roam", "ayr_woodmere": "ayr_woodmere",
}

def base(n):
    return re.sub(r"\s*[\|\[(].*$", "", str(n)).strip().lower()

def norm(s):
    return re.sub(r"[^a-z0-9 ]", " ", str(s).lower()).strip()

def strip_brand(s):
    return re.sub(r"^[^-|]+[-|]\s*", "", str(s))

def num(x):
    if x in (None, "", "—"):
        return 0.0
    try:
        return round(float(x), 3)
    except (TypeError, ValueError):
        return 0.0

def pull_date(wb):
    """Menu pull date from the Notes sheet — same logic as build-strain-data.py.
    Combined line ('...pulled 2026-06-07') OR 2-column ('Pull date' | '2026-06-08')."""
    if "Notes" not in wb.sheetnames:
        return ""
    rows = [" ".join(str(c) for c in r if c is not None) for r in wb["Notes"].iter_rows(values_only=True)]
    for line in rows:  # prefer a line that mentions pulling
        if "pull" in line.lower():
            m = re.search(r"(\d{4}-\d{2}-\d{2})", line)
            if m:
                return m.group(1)
    for line in rows:  # fallback: any date in the Notes sheet
        m = re.search(r"(\d{4}-\d{2}-\d{2})", line)
        if m:
            return m.group(1)
    return ""

def load_flower():
    """cultivar base-name -> richest-terpene representative product; plus brand-token set."""
    d = json.loads(FLOWER.read_text())
    reps, brands = {}, set()
    for p in d["products"]:
        for tok in re.split(r"[|/]", str(p.get("brand", ""))):
            tok = tok.strip().lower()
            if len(tok) >= 4:
                brands.add(tok)
    for p in d["products"]:
        b = base(p["name"])
        if not b:
            continue
        terps = sum((p.get(k, 0) or 0) for k in TERP_KEYS)
        if b not in reps or terps > reps[b][1]:
            reps[b] = (p, terps)
    reps = {k: v[0] for k, v in reps.items() if k not in brands}  # drop brand-name collisions
    store_meta = {x["key"]: x for x in d["dispensaries"]}
    return reps, brands, store_meta

def match_cultivar(text, reps, brands):
    if not text:
        return None
    cand = sorted((c for c in reps if len(c) >= 5), key=len, reverse=True)
    for t in (norm(strip_brand(text)), norm(text)):
        if t in reps:
            return reps[t]
        for c in cand:
            if c in brands:
                continue
            if re.search(r"\b" + re.escape(c) + r"\b", t):
                return reps[c]
    return None

def lead_terp(terps):
    best, bv = None, 0
    pairs = {**terps, "pinene": terps.get("apinene", 0) + terps.get("bpinene", 0)}
    for k in ["myrcene", "limonene", "caryophyllene", "linalool", "humulene", "pinene", "bisabolol"]:
        if pairs.get(k, 0) > bv:
            best, bv = k, pairs[k]
    return best if bv > 0 else None

def main():
    try:
        import openpyxl
    except ImportError:
        sys.exit("pip install openpyxl")
    reps, brands, store_meta = load_flower()
    products, stores = [], {}
    dates = []  # menu pull dates across all edible workbooks -> feed "updated"
    bad_coa = 0  # COA terpene rows rejected for implausible units (mg/g or mg, not % w/w)
    for f in sorted(SRC.glob("*_edibles.xlsx")):
        stem = f.stem.replace("_edibles", "")
        key = STORE_KEY.get(stem, stem)
        meta = store_meta.get(key, {"key": key, "label": key, "location": "", "state": "", "color": "#777"})
        wb = openpyxl.load_workbook(f, data_only=True)
        dates.append(pull_date(wb))
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        H = {h: i for i, h in enumerate(rows[0]) if h is not None}
        def cell(r, col):
            i = H.get(col)
            return r[i] if i is not None and i < len(r) else None
        n = prof = 0
        for r in rows[1:]:
            name = cell(r, "Product")
            if not name:
                continue
            fid = str(cell(r, "Fidelity") or "").strip().upper()
            coa = str(cell(r, "Terps published") or "").lower() == "yes"
            strain, src = cell(r, "Strain"), cell(r, "Source Cultivar")
            terps, profile, matched = None, "none", None
            if coa:  # product publishes its own terpene COA
                t = {k: num(cell(r, lbl)) for k, lbl in zip(TERP_KEYS,
                     ["Beta Myrcene", "Limonene", "Beta Caryophyllene", "Linalool",
                      "Humulene", "Alpha Pinene", "Beta Pinene", "Bisabolol"])}
                tot = sum(t.values())
                # COA terps MUST be plausible % w/w. Several cowork COAs arrived in mg/g or mg
                # (e.g. caryophyllene 961, linalool 204 — impossible as a percent). Reject those
                # untrustworthy units rather than display a fabricated %; the row then falls
                # through to cultivar-inherit. Real flower tops out ~5% total in our data.
                if 0 < tot <= 8 and max(t.values()) <= 5:
                    terps, profile = t, "coa"
                elif tot > 0:
                    bad_coa += 1
            if terps is None and fid in ("A", "B"):  # inherit from a matched cultivar
                fl = match_cultivar(src, reps, brands) or match_cultivar(strain, reps, brands)
                if fl:
                    terps = {k: (fl.get(k, 0) or 0) for k in TERP_KEYS}
                    profile, matched = "inherited", fl["name"]
            rec = {
                "name": str(name), "brand": str(cell(r, "Brand") or ""),
                "form": str(cell(r, "Form") or ""), "dispensary": key, "state": meta.get("state", ""),
                "extractType": str(cell(r, "Extract Type") or "unknown"),
                "fidelity": fid or "C",
                "thcMg": num(cell(r, "THC mg")), "thcMgPkg": num(cell(r, "THC mg pkg")),
                "cbdMg": num(cell(r, "CBD mg")), "ratio": str(cell(r, "Ratio") or ""),
                "other": str(cell(r, "Other cannabinoids") or ""),
                "fastActing": str(cell(r, "Fast-acting") or "").lower() == "yes",
                "cardiac": str(cell(r, "Cardiac Lane") or "").strip(),
                "price": str(cell(r, "Price") or ""), "size": str(cell(r, "Size") or ""),
                "profile": profile, "matchedCultivar": matched,
            }
            if terps:
                rec.update(terps)
                rec["leadTerp"] = lead_terp(terps)
                prof += 1
            products.append(rec)
            n += 1
        tierA = sum(1 for p in products if p["dispensary"] == key and p["fidelity"] == "A")
        stores[key] = {k: meta.get(k) for k in ("key", "label", "location", "state", "color")} | \
                      {"count": n, "tierA": tierA, "profileable": prof}
        print(f"  {key:16} {n:>4} edibles | tier-A {tierA:>3} | profileable {prof:>2}")

    total = len(products)
    distillate = sum(1 for p in products if p["extractType"] in ("distillate", "unknown"))
    profileable = sum(1 for p in products if p["profile"] != "none")
    payload = {
        "updated": max([d for d in dates if d] or [""]),
        "count": total,
        "stores": list(stores.values()),
        "stats": {
            "total": total, "tierA": sum(1 for p in products if p["fidelity"] == "A"),
            "distillate": distillate, "profileable": profileable,
            "coa": sum(1 for p in products if p["profile"] == "coa"),
            "inherited": sum(1 for p in products if p["profile"] == "inherited"),
            "cardiac": sum(1 for p in products if p["cardiac"]),
        },
        "products": products,
    }
    OUT.write_text(json.dumps(payload, indent=1))
    s = payload["stats"]
    print(f"\nwrote {OUT}")
    print(f"  {total} edibles | {s['tierA']} tier-A | {distillate} distillate ({100*distillate//total}% wall) | "
          f"{profileable} profileable ({s['coa']} COA + {s['inherited']} inherited) | {s['cardiac']} cardiac-lane")
    if bad_coa:
        print(f"  ⚠️  rejected {bad_coa} COA terpene row(s) for implausible units (>8% total or >5% single — likely mg/g or mg, not %)")

if __name__ == "__main__":
    main()
