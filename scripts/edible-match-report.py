#!/usr/bin/env python3
"""edible-match-report.py — the REAL size of the matchable edible catalog.

cowork's key insight: raw tier-A counts overstate reality, because most tier-A
"strains" are FLAVOR names ("Watermelon Smash"), not cultivars. The number that
matters is the CULTIVAR-MATCH RATE: of terp-true edibles that name a strain, how
many actually join a real cultivar in data/strain-terpenes.json (so they can
inherit a terpene profile).

This script computes that, with two guards learned from the Terrasana/URB pilot:
  1. BRAND-COLLISION guard — never "match" on a flower BRAND token (e.g.
     "Peninsula Gardens" is a brand, not a cultivar). Brands are excluded.
  2. Phrase match requires a real multi-char cultivar to appear as a whole phrase.

Usage:  python3 scripts/edible-match-report.py            # all *_edibles.xlsx in /terrasana
        python3 scripts/edible-match-report.py terrasana  # one store key
"""
import json, re, sys
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
SRC = REPO / "terrasana"

def base(n):  # cultivar minus size/format suffix
    return re.sub(r"\s*[\|\[(].*$", "", str(n)).strip().lower()

def load_flower():
    d = json.loads((REPO / "data" / "strain-terpenes.json").read_text())
    cultivars, brands = {}, set()
    for p in d["products"]:
        b = base(p["name"])
        if b:
            cultivars.setdefault(b, p["name"])
        for tok in re.split(r"[|/]", str(p.get("brand", ""))):
            tok = tok.strip().lower()
            if len(tok) >= 4:
                brands.add(tok)
    # a "base name" that is really just a brand is NOT a cultivar to match on
    cultivars = {k: v for k, v in cultivars.items() if k not in brands}
    return cultivars, brands

def norm(s):
    return re.sub(r"[^a-z0-9 ]", " ", str(s).lower()).strip()

def strip_brand(s):  # drop a leading "Brand - " / "Brand | " prefix cowork leaves in Strain
    return re.sub(r"^[^-|]+[-|]\s*", "", str(s))

def match(strain, cultivars, brands):
    cand = sorted((c for c in cultivars if len(c) >= 5), key=len, reverse=True)
    for txt in (norm(strip_brand(strain)), norm(strain)):
        if txt in cultivars:
            return cultivars[txt]
        for c in cand:
            if c in brands:
                continue
            if re.search(r"\b" + re.escape(c) + r"\b", txt):
                return cultivars[c]
    return None

def main():
    try:
        import openpyxl
    except ImportError:
        sys.exit("pip install openpyxl")
    cultivars, brands = load_flower()
    only = sys.argv[1] if len(sys.argv) > 1 else None
    files = sorted(SRC.glob("*_edibles.xlsx"))
    grand = {"tot": 0, "A": 0, "coa": 0, "named": 0, "matched": 0}
    print(f"{'store':24} {'edibles':>7} {'tierA':>6} {'COA':>4} {'named':>6} {'MATCHED':>8} {'rate':>6}")
    for f in files:
        key = f.stem.replace("_edibles", "")
        if only and not key.startswith(only):
            continue
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        H = {h: i for i, h in enumerate(rows[0])}
        A = coa = named = matched = 0
        hits = []
        for r in rows[1:]:
            fid = str(r[H.get("Fidelity")] or "").strip().upper()
            strain = r[H.get("Strain")]
            if str(r[H.get("Terps published")] or "").lower() == "yes":
                coa += 1
            if fid == "A":
                A += 1
            if fid in ("A", "B") and strain:
                named += 1
                m = match(strain, cultivars, brands)
                if m:
                    matched += 1
                    hits.append((str(strain)[:34], m))
        tot = len(rows) - 1
        # profileable = published COA  OR  genuine cultivar join
        prof = coa + matched
        rate = f"{100*matched//max(named,1)}%"
        print(f"{key:24} {tot:>7} {A:>6} {coa:>4} {named:>6} {matched:>8} {rate:>6}   (profileable≈{prof})")
        for es, fl in hits:
            print(f"      ✓ '{es}'  ->  {fl}")
        for k, v in (("tot", tot), ("A", A), ("coa", coa), ("named", named), ("matched", matched)):
            grand[k] += v
    print(f"\nTOTAL: {grand['tot']} edibles | {grand['A']} tier-A | "
          f"{grand['coa']} published-COA | {grand['named']} named | "
          f"{grand['matched']} genuine cultivar joins "
          f"({100*grand['matched']//max(grand['named'],1)}% of named, "
          f"{100*grand['matched']//max(grand['tot'],1)}% of all)")
    print(f"==> Terpene-PROFILEABLE catalog ≈ {grand['coa'] + grand['matched']} edibles "
          f"(COA + cultivar-join). That's the matchable universe.")

if __name__ == "__main__":
    main()
