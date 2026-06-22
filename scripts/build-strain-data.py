#!/usr/bin/env python3
"""
build-strain-data.py — merge the dispensary flower terpene spreadsheets into the
single JSON the /strain-finder page reads.

Data flow:  claude-cowork keeps /terrasana/*.xlsx refreshed -> THIS SCRIPT -> data/strain-terpenes.json -> page fetch

Usage:  python3 scripts/build-strain-data.py
Output: data/strain-terpenes.json  ({updated, dispensaries[], count, products[]})

Each source xlsx has a grid sheet with these columns (Size is optional):
Product, [Size], Brand, Type, THC %, Total Terps %, Beta Myrcene, Limonene,
Beta Caryophyllene, Linalool, Humulene, Alpha Pinene, Beta Pinene, Bisabolol,
Caryophyllene Oxide, Eucalyptol, Nerolidol
Columns are mapped BY HEADER NAME, so a missing Size column is handled gracefully.
"""
import json, os, re, sys
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent
SRC_DIR = Path(os.environ.get("TERRASANA_DIR", str(REPO / "terrasana")))
OUT = REPO / "data" / "strain-terpenes.json"

# Each dispensary the coworker keeps refreshed in SRC_DIR.
SOURCES = [
    {"key": "terrasana", "label": "Terrasana", "location": "Cleveland, OH", "state": "OH",
     "color": "#1a8754", "url": "https://terrasanacannabisco.com/cleveland-medical-menu/",
     "file": "terrasana_cleveland_flower_terpenes.xlsx"},
    {"key": "dacut", "label": "Dacut", "location": "Monroe, MI", "state": "MI",
     "color": "#2471a3", "url": "https://dacut.com/",
     "file": "dacut_monroe_flower_terpenes.xlsx"},
    {"key": "rise", "label": "RISE", "location": "Cleveland (Pearl Rd), OH", "state": "OH",
     "color": "#e67e22", "url": "https://oh.risecannabis.com/dispensaries/ohio/cleveland-pearl-rd/6013/recreational-menu/",
     "file": "rise_cleveland_pearl_flower_terpenes.xlsx"},
    {"key": "story_med", "label": "Story (Med)", "location": "Cleveland (Brookpark), OH", "state": "OH",
     "color": "#8e44ad", "url": "https://storycannabis.com/categories/flower/?retailer_id=9bd46ef6-4494-4986-b54b-42c191f26db2&menu_type=MEDICAL&categories=FLOWER",
     "merge_key": "story", "file": "story_cleveland_med_flower_terpenes.xlsx"},
    {"key": "story_rec", "label": "Story (Rec)", "location": "Cleveland (Brookpark), OH", "state": "OH",
     "color": "#c0392b", "url": "https://storycannabis.com/categories/flower/?retailer_id=9bd46ef6-4494-4986-b54b-42c191f26db2&menu_type=RECREATIONAL&categories=FLOWER",
     "merge_key": "story", "file": "story_cleveland_rec_flower_terpenes.xlsx"},
    {"key": "urb", "label": "URB", "location": "Monroe, MI", "state": "MI",
     "color": "#16a085", "url": "https://dutchie.com/stores/urb-cannabis-monroe/products/flower",
     "file": "urb_cannabis_monroe_flower_terpenes.xlsx"},
    {"key": "joyology", "label": "Joyology", "location": "Monroe, MI", "state": "MI",
     "color": "#d35400", "url": "https://dutchie.com/stores/grams-club-monroe-rec/products/flower",
     "file": "joyology_monroe_flower_terpenes.xlsx"},
    {"key": "puff", "label": "PUFF", "location": "Monroe, MI", "state": "MI",
     "color": "#2980b9", "url": "https://dutchie.com/stores/puff-monroe-rec/products/flower",
     "file": "puff_monroe_flower_terpenes.xlsx"},
    {"key": "pure", "label": "PURE", "location": "Monroe, MI", "state": "MI",
     "color": "#27ae60", "url": "https://dutchie.com/stores/pure-dixie-monroe-rec/products/flower",
     "file": "pure_monroe_flower_terpenes.xlsx"},
    {"key": "klutch", "label": "Klutch", "location": "Cleveland, OH", "state": "OH",
     "color": "#9b59b6", "url": "https://dutchie.com/stores/klutch-cannabis-cleveland/products/flower",
     "file": "klutch_cleveland_flower_terpenes.xlsx"},
    {"key": "botanist_solon", "label": "The Botanist", "location": "Solon, OH", "state": "OH",
     "color": "#2ecc71", "url": "https://shopbotanist.com/stores/solon-menu/products/flower",
     "file": "botanist_solon_flower_terpenes.xlsx"},
    {"key": "landing", "label": "The Landing", "location": "Cleveland, OH", "state": "OH",
     "color": "#e74c3c", "url": "https://cleveland-menu.thelandingdispensaries.com/stores/cleveland-ohio/products/flower",
     "file": "landing_cleveland_flower_terpenes.xlsx"},
    {"key": "amplify", "label": "Amplify", "location": "Cleveland Heights, OH", "state": "OH",
     "color": "#f39c12", "url": "https://amplifydispensary.com/stores/amplify-coventry/products/flower",
     "file": "amplify_cleveland_hts_flower_terpenes.xlsx"},
    {"key": "forest_lakewood", "label": "The Forest", "location": "Lakewood, OH", "state": "OH",
     "color": "#1abc9c", "url": "https://theforestdispensary.com/stores/the-forest-lakewood/products/flower",
     "file": "forest_lakewood_flower_terpenes.xlsx"},
    {"key": "roam", "label": "ROAM", "location": "Seven Hills, OH", "state": "OH",
     "color": "#e67e22", "url": "https://roamdispensary.com/adult-use/",
     "file": "roam_seven_hills_flower_terpenes.xlsx"},
    {"key": "ayr_woodmere", "label": "AYR", "location": "Woodmere, OH", "state": "OH",
     "color": "#16a085", "url": "https://ayrdispensaries.com/ohio/woodmere/shop/",
     "file": "ayr_woodmere_flower_terpenes.xlsx"},
]

# Sources sharing a merge_key collapse into one dispensary, deduping strains that
# appear on more than one of their menus (Story's Med and Rec menus are ~99% identical
# — same flower, same shelf — so two chips and duplicate result cards add only noise).
MERGE_META = {
    "story": {"key": "story", "label": "Story", "location": "Cleveland (Brookpark), OH", "state": "OH",
              "color": "#8e44ad",
              "url": "https://storycannabis.com/categories/flower/?retailer_id=9bd46ef6-4494-4986-b54b-42c191f26db2&categories=FLOWER"},
}

# header label (lowercased) -> output field
COLMAP = {
    "product": "name", "size": "size", "brand": "brand", "type": "type",
    "thc %": "thc", "total terps %": "terps",
    "beta myrcene": "myrcene", "limonene": "limonene", "beta caryophyllene": "caryophyllene",
    "linalool": "linalool", "humulene": "humulene", "alpha pinene": "apinene",
    "beta pinene": "bpinene", "bisabolol": "bisabolol",
    "caryophyllene oxide": "caryophylleneoxide", "eucalyptol": "eucalyptol", "nerolidol": "nerolidol",
}
NUM_FIELDS = {"thc", "terps", "myrcene", "limonene", "caryophyllene", "linalool", "humulene",
              "apinene", "bpinene", "bisabolol", "caryophylleneoxide", "eucalyptol", "nerolidol"}

# Not flower: moonrocks and infused prepacks are nug coated in concentrate + kief, so their
# THC (40-62%) dwarfs any real flower and would top the THC "intensity" sort with a number no
# bud can hit. Their terpenes are fine, but the category isn't what this flower tool ranks.
# Pattern is deliberately narrow so strain NAMES that merely contain a word (Hash Burger,
# Hash Queen, Diamond Bar) stay in.
EXCLUDE_NAME = re.compile(r"infused|moon ?rocks?", re.I)

def num(x):
    if x is None or x == "—" or x == "":
        return 0.0
    try:
        return round(float(x), 3)
    except (TypeError, ValueError):
        return 0.0

def pull_date(wb):
    """Find the menu pull date in the Notes sheet, tolerant of layout:
    combined line ('...pulled 2026-06-07') OR 2-column ('Pull date' | '2026-06-08')."""
    if "Notes" not in wb.sheetnames:
        return ""
    rows = [" ".join(str(c) for c in r if c is not None) for r in wb["Notes"].iter_rows(values_only=True)]
    for line in rows:  # prefer a line that mentions pulling
        if "pull" in line.lower():
            m = re.search(r"(\d{4}-\d{2}-\d{2})", line)
            if m:
                return m.group(1)
    for line in rows:  # fallback: any date in the Notes sheet
        m = re.search(r"(\d{4}-\d{2}-\d{2})", line)
        if m:
            return m.group(1)
    return ""

def main():
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl required:  pip install openpyxl")

    products, dispensaries, dates = [], [], []
    merge_groups = {}  # merge_key -> {"dates": [...], "seen": set()} for cross-menu dedup
    excluded = 0       # infused / moonrock rows skipped (not flower)
    for src in SOURCES:
        path = SRC_DIR / src["file"]
        if not path.exists():
            print(f"  skip {src['key']}: {path} not found")
            continue
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        header = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
        idx = {COLMAP[h]: i for i, h in enumerate(header) if h in COLMAP}
        updated = pull_date(wb)
        dates.append(updated)
        mk = src.get("merge_key")
        disp_key = mk or src["key"]
        mg = merge_groups.setdefault(mk, {"dates": [], "seen": set()}) if mk else None
        n = 0
        for r in rows[1:]:
            if not r or idx.get("name") is None or not r[idx["name"]]:
                continue
            if EXCLUDE_NAME.search(str(r[idx["name"]])):  # moonrocks / infused — not flower
                excluded += 1
                continue
            rec = {"dispensary": disp_key}
            for field, i in idx.items():
                v = r[i] if i < len(r) else None
                rec[field] = num(v) if field in NUM_FIELDS else (v if v is not None else "")
            rec.setdefault("size", "")
            # backfill a missing "Total Terps %" from the individual terpene columns
            # (some labs, e.g. Firelands at The Landing, leave the total blank but list the parts)
            if not rec.get("terps"):
                rec["terps"] = round(sum(rec.get(k, 0) or 0 for k in
                    ("myrcene", "limonene", "caryophyllene", "linalool", "humulene",
                     "apinene", "bpinene", "bisabolol", "caryophylleneoxide", "eucalyptol", "nerolidol")), 3)
            if mg is not None:  # dedup across the merged menus by name+size
                dk = (str(rec.get("name", "")).strip().lower(), str(rec.get("size", "")).strip().lower())
                if dk in mg["seen"]:
                    continue
                mg["seen"].add(dk)
            products.append(rec)
            n += 1
        if mk:
            mg["dates"].append(updated)
            print(f"  {src['key']}: {n} new products into '{mk}' (updated {updated or 'unknown'})")
        else:
            dispensaries.append({k: src[k] for k in ("key", "label", "location", "state", "color", "url")} | {"count": n, "updated": updated})
            print(f"  {src['key']}: {n} products (updated {updated or 'unknown'})")

    # emit one dispensary entry per merge group (after the standalone ones, preserving bar order)
    for mk, mg in merge_groups.items():
        meta = MERGE_META[mk]
        cnt = sum(1 for p in products if p["dispensary"] == mk)
        upd = max([d for d in mg["dates"] if d] or [""])
        dispensaries.append({k: meta[k] for k in ("key", "label", "location", "state", "color", "url")} | {"count": cnt, "updated": upd})
        print(f"  -> merged '{mk}': {cnt} unique products (updated {upd or 'unknown'})")

    payload = {
        "updated": max([d for d in dates if d] or [""]),
        "dispensaries": dispensaries,
        "count": len(products),
        "products": products,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, indent=1))
    print(f"wrote {OUT}  ({len(products)} products across {len(dispensaries)} dispensaries; {excluded} infused/moonrock rows excluded)")

if __name__ == "__main__":
    main()
